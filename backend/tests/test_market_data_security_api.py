from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

pytest.importorskip("httpx", reason="fastapi TestClient requires httpx")
from fastapi.testclient import TestClient

from option_risk.api import app
from option_risk.data.market_data_sessions import create_market_data_session, store_market_data_file


def _normal_xlsx_bytes(row_count: int) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row_index in range(row_count):
        sheet.append([row_index])
    return _save_workbook(workbook)


def _write_only_xlsx_bytes(row_count: int) -> bytes:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet()
    for row_index in range(row_count):
        sheet.append([row_index])
    return _save_workbook(workbook)


def _empty_xlsx_bytes() -> bytes:
    return _save_workbook(Workbook())


def _save_workbook(workbook: Workbook) -> bytes:
    buf = BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _upload_curve_discount(client: TestClient, payload: bytes):
    return client.post(
        "/market-data/upload",
        files={
            "file": (
                "curveDiscount.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_market_data_upload_allows_normal_xlsx_at_row_limit(monkeypatch, tmp_path):
    monkeypatch.setenv("OPTION_RISK_MARKET_SESSION_ROOT", str(tmp_path / "sessions"))
    client = TestClient(app)

    resp = _upload_curve_discount(client, _normal_xlsx_bytes(5000))

    assert resp.status_code == 200
    assert resp.json()["files"][0]["kind"] == "curve_discount"


@pytest.mark.parametrize(
    "payload_factory",
    [
        _normal_xlsx_bytes,
        _write_only_xlsx_bytes,
    ],
)
def test_market_data_upload_rejects_xlsx_over_row_limit(monkeypatch, tmp_path, payload_factory):
    monkeypatch.setenv("OPTION_RISK_MARKET_SESSION_ROOT", str(tmp_path / "sessions"))
    client = TestClient(app)

    resp = _upload_curve_discount(client, payload_factory(5001))

    assert resp.status_code == 400
    assert "превышает лимит строк XLSX" in resp.json()["message"]
    assert "максимум 5000" in resp.json()["message"]


@pytest.mark.parametrize(
    ("payload_factory", "row_count"),
    [
        (_empty_xlsx_bytes, None),
        (_normal_xlsx_bytes, 1),
    ],
    ids=["empty", "single-row"],
)
def test_market_data_upload_allows_empty_and_single_row_xlsx(monkeypatch, tmp_path, payload_factory, row_count):
    monkeypatch.setenv("OPTION_RISK_MARKET_SESSION_ROOT", str(tmp_path / "sessions"))
    client = TestClient(app)
    payload = payload_factory() if row_count is None else payload_factory(row_count)

    resp = _upload_curve_discount(client, payload)

    assert resp.status_code == 200


def test_market_data_upload_rejects_corrupted_xlsx_with_clean_error(monkeypatch, tmp_path):
    monkeypatch.setenv("OPTION_RISK_MARKET_SESSION_ROOT", str(tmp_path / "sessions"))
    client = TestClient(app)

    resp = _upload_curve_discount(client, b"not an xlsx archive")

    assert resp.status_code == 400
    payload = resp.json()
    assert "корректным XLSX" in payload["message"]
    assert "Traceback" not in str(payload)


@pytest.mark.parametrize(
    "payload_factory",
    [
        _normal_xlsx_bytes,
        _write_only_xlsx_bytes,
    ],
)
def test_store_market_data_file_raises_value_error_for_xlsx_over_row_limit(monkeypatch, tmp_path, payload_factory):
    monkeypatch.setenv("OPTION_RISK_MARKET_SESSION_ROOT", str(tmp_path / "sessions"))
    session_id = create_market_data_session()

    with pytest.raises(ValueError, match="превышает лимит строк XLSX"):
        store_market_data_file(session_id, "curveDiscount.xlsx", payload_factory(5001))
