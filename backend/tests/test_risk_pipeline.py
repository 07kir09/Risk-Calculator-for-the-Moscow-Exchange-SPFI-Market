import datetime as dt
import math
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

from option_risk.data.models import MarketScenario, OptionPosition, OptionStyle, OptionType, Portfolio
from option_risk.risk.limits import check_limits
from option_risk.risk.pipeline import CalculationConfig, run_calculation
from option_risk.risk.correlations import correlation_matrix
from option_risk.risk.portfolio import scenario_pnl
from option_risk.risk.var_es import historical_var
from option_risk.pricing import mc_price
from option_risk.pricing.market import MarketDataContext


def make_position(**kwargs) -> OptionPosition:
    base = dict(
        position_id="test",
        instrument_type="option",
        option_type=OptionType.CALL,
        style=OptionStyle.EUROPEAN,
        quantity=1,
        notional=1.0,
        underlying_symbol="TEST",
        underlying_price=100.0,
        strike=100.0,
        volatility=0.2,
        maturity_date=dt.date(2026, 1, 1),
        valuation_date=dt.date(2025, 1, 1),
        risk_free_rate=0.05,
        dividend_yield=0.0,
        currency="RUB",
    )
    base.update(kwargs)
    return OptionPosition(**base)


def _load_workbook_positions(row_numbers: list[int]) -> list[OptionPosition]:
    workbook_path = Path(__file__).resolve().parents[2] / "Datasets" / "portfolio_large_1000.xlsx"
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    positions: list[OptionPosition] = []
    for row_number in row_numbers:
        raw_row = next(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
        positions.append(OptionPosition(**dict(zip(headers, raw_row))))
    return positions


def _manual_forward_value(position: OptionPosition, fx_rate: float) -> float:
    tenor_years = (position.maturity_date - position.valuation_date).days / 365.0
    discounted = math.exp(-position.risk_free_rate * tenor_years)
    return position.quantity * position.notional * (position.underlying_price - position.strike) * discounted * fx_rate


def test_check_limits_var_breaches_when_above_limit():
    limits = {"var_hist": 1000.0}
    metrics = {"var_hist": 1200.0}
    res = check_limits(metrics, limits)
    assert res == [("var_hist", 1200.0, 1000.0, True)]


def test_check_limits_pnl_breaches_when_below_minus_limit():
    limits = {"daily_pnl": 1000.0}
    metrics = {"daily_pnl": -1500.0}
    res = check_limits(metrics, limits)
    assert res == [("daily_pnl", -1500.0, 1000.0, True)]


def test_run_calculation_end_to_end():
    portfolio = Portfolio(positions=[make_position(position_id="a"), make_position(position_id="b", strike=90)])
    scenarios = [
        MarketScenario(scenario_id="s1", underlying_shift=-0.05, volatility_shift=0.1, rate_shift=0.0),
        MarketScenario(scenario_id="s2", underlying_shift=0.03, volatility_shift=-0.05, rate_shift=0.0005),
        MarketScenario(scenario_id="s3", underlying_shift=0.0, volatility_shift=0.0, rate_shift=-0.0005),
    ]
    limits_cfg = {"var_hist": 1.0, "stress": {"s1": 1.0}}
    cfg = CalculationConfig(calc_sensitivities=True, calc_var_es=True, calc_stress=True, calc_margin_capital=True, alpha=0.95)
    result = run_calculation(portfolio, scenarios, limits_cfg=limits_cfg, config=cfg)

    assert np.isfinite(result.base_value)
    assert result.greeks is not None
    assert "dv01" in result.greeks
    assert result.stress is not None and len(result.stress) == len(scenarios)
    assert result.var_hist is not None and result.var_hist >= 0
    assert result.es_hist is not None and result.es_hist >= result.var_hist
    assert result.lc_var is not None and result.lc_var >= result.var_hist
    assert result.limits is not None
    assert result.initial_margin is not None


def test_run_calculation_respects_flags():
    portfolio = Portfolio(positions=[make_position()])
    scenarios = [MarketScenario(scenario_id="s1", underlying_shift=0.0, volatility_shift=0.0, rate_shift=0.0)]
    cfg = CalculationConfig(calc_sensitivities=False, calc_var_es=False, calc_stress=False, calc_margin_capital=False)
    result = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)
    assert result.greeks is None
    assert result.var_hist is None
    assert result.es_hist is None
    assert result.stress is None
    assert result.capital is None


def test_run_calculation_does_not_report_zero_limits_for_skipped_var_metrics():
    portfolio = Portfolio(positions=[make_position()])
    scenarios = [MarketScenario(scenario_id="s1", underlying_shift=0.0, volatility_shift=0.0, rate_shift=0.0)]
    limits_cfg = {"var_hist": 1.0, "es_hist": 1.0, "lc_var": 1.0, "stress": {"s1": 1.0}}
    cfg = CalculationConfig(calc_sensitivities=False, calc_var_es=False, calc_stress=True, calc_margin_capital=False)

    result = run_calculation(portfolio, scenarios, limits_cfg=limits_cfg, config=cfg)

    assert result.var_hist is None
    assert result.es_hist is None
    assert result.lc_var is None
    assert result.limits == []
    assert result.stress is not None


def test_correlation_matrix_requires_two_scenarios():
    portfolio = Portfolio(positions=[make_position(position_id="a"), make_position(position_id="b", strike=90)])
    scenarios = [MarketScenario(scenario_id="only", underlying_shift=0.01, volatility_shift=0.0, rate_shift=0.0)]
    with pytest.raises(ValueError):
        correlation_matrix(portfolio, scenarios)


def test_monte_carlo_is_deterministic_with_seed():
    position = make_position()
    p1 = mc_price(position, n_paths=5000, seed=123)
    p2 = mc_price(position, n_paths=5000, seed=123)
    assert p1 == p2


def test_run_calculation_sanitizes_nan_correlations():
    portfolio = Portfolio(
        positions=[
            make_position(
                position_id="f1",
                instrument_type="forward",
                volatility=0.0,
                underlying_price=100.0,
                strike=100.0,
                notional=1.0,
            ),
            make_position(
                position_id="f2",
                instrument_type="forward",
                volatility=0.0,
                underlying_price=100.0,
                strike=100.0,
                notional=1.0,
            ),
        ]
    )
    scenarios = [
        MarketScenario(scenario_id="s1", underlying_shift=0.0, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="s2", underlying_shift=0.0, volatility_shift=0.0, rate_shift=0.0),
    ]
    cfg = CalculationConfig(calc_sensitivities=True, calc_var_es=True, calc_stress=True, calc_margin_capital=True, alpha=0.99)
    result = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)

    assert result.correlations is not None
    corr = np.asarray(result.correlations, dtype=np.float64)
    assert np.isfinite(corr).all()
    assert np.allclose(np.diag(corr), 1.0)
    assert any("корреляц" in msg.message.lower() for msg in result.validation_log)


def test_run_calculation_skips_correlations_when_disabled():
    portfolio = Portfolio(positions=[make_position(position_id="a"), make_position(position_id="b", strike=95)])
    scenarios = [
        MarketScenario(scenario_id="s1", underlying_shift=-0.01, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="s2", underlying_shift=0.01, volatility_shift=0.0, rate_shift=0.0),
    ]
    cfg = CalculationConfig(calc_sensitivities=True, calc_var_es=True, calc_stress=True, calc_margin_capital=True, calc_correlations=False)
    result = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)

    assert result.correlations is None


def test_run_calculation_skips_correlations_when_position_count_exceeds_limit():
    positions = [
        make_position(
            position_id=f"f{i}",
            instrument_type="forward",
            volatility=0.0,
            underlying_price=100.0 + (i % 3),
            strike=100.0,
            notional=1.0,
        )
        for i in range(6)
    ]
    portfolio = Portfolio(positions=positions)
    scenarios = [
        MarketScenario(scenario_id="s1", underlying_shift=-0.01, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="s2", underlying_shift=0.01, volatility_shift=0.0, rate_shift=0.0),
    ]
    cfg = CalculationConfig(
        calc_sensitivities=False,
        calc_var_es=True,
        calc_stress=False,
        calc_margin_capital=False,
        calc_correlations=True,
        max_correlation_positions=4,
    )
    result = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)

    assert result.correlations is None
    assert any("корреляц" in msg.message.lower() and "пропущен" in msg.message.lower() for msg in result.validation_log)


def test_run_calculation_truncates_lc_breakdown_rows():
    portfolio = Portfolio(
        positions=[
            make_position(position_id=f"p{i}", liquidity_haircut=0.5 + i * 0.1, notional=100.0 + i * 10.0)
            for i in range(8)
        ]
    )
    scenarios = [MarketScenario(scenario_id="s1", underlying_shift=-0.02, volatility_shift=0.0, rate_shift=0.0)]
    cfg = CalculationConfig(
        calc_sensitivities=False,
        calc_var_es=True,
        calc_stress=False,
        calc_margin_capital=False,
        max_lc_breakdown_rows=3,
    )
    result = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)

    assert result.lc_var_breakdown is not None
    assert len(result.lc_var_breakdown) == 3
    assert any("breakdown" in msg.message.lower() and "top-3" in msg.message.lower() for msg in result.validation_log)


def test_run_calculation_omits_large_pnl_matrix_from_response():
    portfolio = Portfolio(
        positions=[
            make_position(
                position_id=f"f{i}",
                instrument_type="forward",
                volatility=0.0,
                underlying_price=100.0 + i,
                strike=100.0,
                notional=1.0,
            )
            for i in range(6)
        ]
    )
    scenarios = [
        MarketScenario(scenario_id="s1", underlying_shift=-0.01, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="s2", underlying_shift=0.0, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="s3", underlying_shift=0.01, volatility_shift=0.0, rate_shift=0.0),
    ]
    cfg = CalculationConfig(
        calc_sensitivities=False,
        calc_var_es=True,
        calc_stress=False,
        calc_margin_capital=False,
        max_pnl_matrix_cells=10,
    )
    result = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)

    assert result.pnl_matrix is None
    assert any("pnl_matrix" in msg.message.lower() for msg in result.validation_log)


def test_run_calculation_uses_weighted_historical_var_when_probabilities_provided():
    portfolio = Portfolio(
        positions=[
            make_position(
                instrument_type="forward",
                volatility=0.0,
                underlying_price=100.0,
                strike=100.0,
                notional=1.0,
            )
        ]
    )
    scenarios = [
        MarketScenario(
            scenario_id="rare_crash",
            underlying_shift=-0.5,
            volatility_shift=0.0,
            rate_shift=0.0,
            probability=0.01,
        ),
        MarketScenario(
            scenario_id="common_move",
            underlying_shift=-0.01,
            volatility_shift=0.0,
            rate_shift=0.0,
            probability=0.99,
        ),
    ]
    cfg = CalculationConfig(
        calc_sensitivities=False,
        calc_var_es=True,
        calc_stress=False,
        calc_margin_capital=False,
        calc_correlations=False,
        alpha=0.95,
    )
    result = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)

    assert result.pnl_distribution is not None
    unweighted_var = historical_var(result.pnl_distribution, alpha=0.95)
    assert result.var_hist is not None
    assert result.var_hist < unweighted_var


def test_run_calculation_requires_fx_for_foreign_currency_positions():
    positions = _load_workbook_positions([704, 705, 706, 707, 708])
    portfolio = Portfolio(positions=positions)
    cfg = CalculationConfig(
        base_currency="RUB",
        fx_rates={},
        calc_sensitivities=False,
        calc_var_es=False,
        calc_stress=False,
        calc_margin_capital=False,
        calc_correlations=False,
    )

    with pytest.raises(ValueError, match=r"USD.*USD/RUB"):
        run_calculation(portfolio, [], limits_cfg=None, config=cfg)


def test_variation_margin_prefers_named_base_scenario():
    portfolio = Portfolio(
        positions=[
            make_position(
                position_id="vm_fwd",
                instrument_type="forward",
                underlying_price=100.0,
                strike=90.0,
                volatility=0.0,
                quantity=1.0,
                notional=1.0,
                risk_free_rate=0.05,
            )
        ]
    )
    scenarios = [
        MarketScenario(scenario_id="stress_up", underlying_shift=0.15, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="base", underlying_shift=0.0, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="stress_down", underlying_shift=-0.10, volatility_shift=0.0, rate_shift=0.0),
    ]
    reordered = [scenarios[2], scenarios[0], scenarios[1]]
    cfg = CalculationConfig(
        calc_sensitivities=False,
        calc_var_es=True,
        calc_stress=False,
        calc_margin_capital=True,
        calc_correlations=False,
    )

    result_a = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)
    result_b = run_calculation(portfolio, reordered, limits_cfg=None, config=cfg)

    reference = next(item for item in scenarios if item.scenario_id == "base")
    expected = scenario_pnl(portfolio, reference)

    assert result_a.variation_margin is not None
    assert result_b.variation_margin is not None
    assert result_a.variation_margin == pytest.approx(result_b.variation_margin)
    assert result_a.variation_margin == pytest.approx(expected)


def test_variation_margin_falls_back_to_last_scenario_name():
    portfolio = Portfolio(
        positions=[
            make_position(
                position_id="vm_fwd_last",
                instrument_type="forward",
                underlying_price=100.0,
                strike=90.0,
                volatility=0.0,
                quantity=1.0,
                notional=1.0,
                risk_free_rate=0.05,
            )
        ]
    )
    scenarios = [
        MarketScenario(scenario_id="stress_down", underlying_shift=-0.10, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="last_scenario", underlying_shift=0.25, volatility_shift=0.0, rate_shift=0.0),
        MarketScenario(scenario_id="stress_up", underlying_shift=0.15, volatility_shift=0.0, rate_shift=0.0),
    ]
    reordered = [scenarios[2], scenarios[0], scenarios[1]]
    cfg = CalculationConfig(
        calc_sensitivities=False,
        calc_var_es=True,
        calc_stress=False,
        calc_margin_capital=True,
        calc_correlations=False,
    )

    result_a = run_calculation(portfolio, scenarios, limits_cfg=None, config=cfg)
    result_b = run_calculation(portfolio, reordered, limits_cfg=None, config=cfg)

    reference = next(item for item in scenarios if item.scenario_id == "last_scenario")
    expected = scenario_pnl(portfolio, reference)

    assert result_a.variation_margin is not None
    assert result_b.variation_margin is not None
    assert result_a.variation_margin == pytest.approx(result_b.variation_margin)
    assert result_a.variation_margin == pytest.approx(expected)


def test_run_calculation_with_full_fx_matches_manual_five_row_check():
    positions = _load_workbook_positions([704, 705, 706, 707, 708])
    portfolio = Portfolio(positions=positions)
    cfg = CalculationConfig(
        base_currency="RUB",
        fx_rates={"USD": 90.0},
        calc_sensitivities=False,
        calc_var_es=False,
        calc_stress=False,
        calc_margin_capital=False,
        calc_correlations=False,
    )

    result = run_calculation(portfolio, [], limits_cfg=None, config=cfg)
    expected = sum(_manual_forward_value(position, 90.0 if position.currency == "USD" else 1.0) for position in positions)

    assert result.base_value == pytest.approx(expected, rel=1e-12, abs=1e-9)
    assert result.fx_warning is None


def test_run_calculation_uses_market_context_fx_when_request_fx_missing():
    position = make_position(
        position_id="usd_forward",
        instrument_type="forward",
        currency="USD",
        underlying_price=100.0,
        strike=90.0,
        volatility=0.0,
        quantity=1.0,
        notional=1.0,
        risk_free_rate=0.05,
    )
    portfolio = Portfolio(positions=[position])
    market = MarketDataContext(
        discount_curves={},
        forward_curves={},
        fx_spots={"USD": 90.0},
        base_currency="RUB",
    )
    cfg = CalculationConfig(
        base_currency="RUB",
        fx_rates=None,
        calc_sensitivities=False,
        calc_var_es=False,
        calc_stress=False,
        calc_margin_capital=False,
        calc_correlations=False,
    )

    result = run_calculation(portfolio, [], limits_cfg=None, config=cfg, market=market)

    assert result.base_value == pytest.approx(_manual_forward_value(position, 90.0), rel=1e-12, abs=1e-9)
    assert result.fx_warning is None
