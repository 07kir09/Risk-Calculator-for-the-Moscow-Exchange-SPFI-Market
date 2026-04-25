"""Session storage for incremental market-data uploads from UI/API."""
from __future__ import annotations

from io import BytesIO
import json
import logging
import os
import re
import shutil
import tempfile
import unicodedata
import uuid
from zipfile import BadZipFile
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Literal
from urllib.parse import unquote

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .market_data import MarketDataBundle, load_market_data_bundle_from_directory
from .live_market_data import sync_live_market_data_to_directory
from .validation import ValidationMessage

MarketDataFileKind = Literal["curve_discount", "curve_forward", "fixing", "calibration", "fx_history"]

logger = logging.getLogger(__name__)

_REQUIRED_DISPLAY_NAMES = ("curveDiscount.xlsx", "curveForward.xlsx", "fixing.xlsx")
_REQUIRED_KINDS: dict[str, MarketDataFileKind] = {
    "curveDiscount.xlsx": "curve_discount",
    "curveForward.xlsx": "curve_forward",
    "fixing.xlsx": "fixing",
}
_SESSION_ID_LENGTH = 32
_SESSION_ID_RE = re.compile(rf"^[A-Za-z0-9_-]{{{_SESSION_ID_LENGTH}}}$")
_SESSION_ID_ERROR = "Некорректный session_id."
_DEFAULT_MAX_XLSX_ROWS = 5000


@dataclass
class UploadedMarketDataFile:
    filename: str
    kind: MarketDataFileKind
    size_bytes: int


@dataclass
class MarketDataSessionSummary:
    session_id: str
    files: list[UploadedMarketDataFile] = field(default_factory=list)
    missing_required_files: list[str] = field(default_factory=list)
    blocking_errors: int = 0
    warnings: int = 0
    ready: bool = False
    validation_log: list[ValidationMessage] = field(default_factory=list)
    counts: dict[str, int] = field(default_factory=dict)
    available_fx_pairs: list[str] = field(default_factory=list)


def validate_market_data_session_id(session_id: str) -> str:
    if not isinstance(session_id, str):
        raise ValueError(_SESSION_ID_ERROR)

    candidate = session_id
    while True:
        decoded = unquote(candidate)
        if decoded == candidate:
            break
        candidate = decoded
    candidate = unicodedata.normalize("NFKC", candidate)
    if not _SESSION_ID_RE.fullmatch(candidate):
        raise ValueError(_SESSION_ID_ERROR)
    return candidate


def _sessions_root() -> Path:
    configured = os.environ.get("OPTION_RISK_MARKET_SESSION_ROOT")
    if configured:
        return Path(configured)
    return Path(tempfile.gettempdir()) / "option_risk_market_data_sessions"


def _default_datasets_dir() -> Path:
    configured = os.environ.get("OPTION_RISK_DEFAULT_DATASETS_DIR")
    if configured:
        return Path(configured)
    repo_root = Path(__file__).resolve().parents[3]
    candidates = [
        repo_root / "datasets" / "Данные для работы",
        repo_root / "datasets",
        repo_root / "Datasets",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _configured_max_xlsx_rows() -> int:
    raw = os.environ.get("OPTION_RISK_MAX_XLSX_ROWS")
    if raw is None:
        return _DEFAULT_MAX_XLSX_ROWS
    try:
        max_rows = int(raw)
    except ValueError as exc:
        raise ValueError("OPTION_RISK_MAX_XLSX_ROWS должен быть положительным целым числом.") from exc
    if max_rows <= 0:
        raise ValueError("OPTION_RISK_MAX_XLSX_ROWS должен быть положительным целым числом.")
    return max_rows


def validate_market_data_xlsx_row_limit(filename: str, source: bytes | Path, *, max_rows: int | None = None) -> None:
    if not filename.lower().endswith(".xlsx"):
        return

    limit = max_rows if max_rows is not None else _configured_max_xlsx_rows()
    workbook_handle = None if isinstance(source, bytes) else source.open("rb")
    workbook_source = BytesIO(source) if isinstance(source, bytes) else workbook_handle
    workbook = None
    try:
        try:
            workbook = load_workbook(workbook_source, read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError) as exc:
            raise ValueError(f"Файл {filename} не является корректным XLSX.") from exc
        for sheet in workbook.worksheets:
            for row_index, _row in enumerate(sheet.iter_rows(), start=1):
                if row_index > limit:
                    raise ValueError(
                        f"Файл {filename} превышает лимит строк XLSX: максимум {limit}, "
                        f"лист {sheet.title} содержит больше строк."
                    )
    finally:
        if workbook is not None:
            workbook.close()
        if workbook_handle is not None:
            workbook_handle.close()


def _validate_xlsx_row_limit(filename: str, source: bytes | Path, *, max_rows: int | None = None) -> None:
    validate_market_data_xlsx_row_limit(filename, source, max_rows=max_rows)


def _session_dir(session_id: str) -> Path:
    return _sessions_root() / validate_market_data_session_id(session_id)


def get_market_data_session_dir(session_id: str) -> Path:
    return _session_dir(session_id)


def create_market_data_session() -> str:
    session_id = uuid.uuid4().hex
    session_dir = _session_dir(session_id)
    session_dir.mkdir(parents=True, exist_ok=True)
    return session_id


def clear_market_data_session(session_id: str) -> None:
    shutil.rmtree(_session_dir(session_id), ignore_errors=True)


def classify_market_data_filename(filename: str) -> MarketDataFileKind | None:
    name = Path(filename).name
    lower = name.lower()
    if lower in {"curvediscount.xlsx", "curvediscount.xls"}:
        return "curve_discount"
    if lower in {"curveforward.xlsx", "curveforward.xls"}:
        return "curve_forward"
    if lower in {"fixing.xlsx", "fixing.xls"}:
        return "fixing"
    if lower.startswith("calibrationinstrument") and lower.endswith((".xlsx", ".xls")):
        return "calibration"
    if lower.startswith("rc_") and lower.endswith((".xlsx", ".xls")):
        return "fx_history"
    if lower.startswith("market_data_") and lower.endswith((".xlsx", ".xls")):
        return "fx_history"
    return None


def _list_session_files(session_dir: Path) -> list[UploadedMarketDataFile]:
    files: list[UploadedMarketDataFile] = []
    if not session_dir.exists():
        return files
    for path in sorted(p for p in session_dir.iterdir() if p.is_file()):
        kind = classify_market_data_filename(path.name)
        if kind is None:
            continue
        files.append(UploadedMarketDataFile(filename=path.name, kind=kind, size_bytes=path.stat().st_size))
    return files


def _counts_from_bundle(bundle: MarketDataBundle) -> dict[str, int]:
    return {
        "discount_curves": len(bundle.discount_curves),
        "forward_curves": len(bundle.forward_curves),
        "fixings": len(bundle.fixings),
        "calibration_instruments": len(bundle.calibration_instruments),
        "fx_history": len(bundle.fx_history),
    }


def _available_fx_pairs_from_bundle(bundle: MarketDataBundle) -> list[str]:
    if bundle.fx_history.empty or "currency_code" not in bundle.fx_history.columns:
        return []

    pairs: set[str] = set()
    for raw_code in bundle.fx_history["currency_code"].dropna().unique():
        code = str(raw_code).strip().upper()
        if not code or code == "UNK":
            continue
        if "/" in code:
            left, right = [part.strip().upper() for part in code.split("/", 1)]
            if len(left) == 3 and len(right) == 3 and left != right:
                pairs.add(f"{left}/{right}")
            continue
        if len(code) == 3 and code != "RUB":
            pairs.add(f"{code}/RUB")
    return sorted(pairs)


def summarize_market_data_session(session_id: str) -> MarketDataSessionSummary:
    session_dir = _session_dir(session_id)
    files = _list_session_files(session_dir)
    if not session_dir.exists() or not files:
        return MarketDataSessionSummary(
            session_id=session_id,
            files=files,
            missing_required_files=list(_REQUIRED_DISPLAY_NAMES),
            counts=_counts_from_bundle(
                MarketDataBundle(
                    fx_history=_empty_frame(),
                    calibration_instruments=_empty_frame(),
                    discount_curves=_empty_frame(),
                    forward_curves=_empty_frame(),
                    fixings=_empty_frame(),
                    validation_log=[],
                )
            ),
            available_fx_pairs=[],
        )

    bundle = load_market_data_bundle_from_directory(session_dir, strict=False)
    available_kinds = {file.kind for file in files}
    missing_required_files = [name for name, kind in _REQUIRED_KINDS.items() if kind not in available_kinds]
    blocking_errors = sum(1 for message in bundle.validation_log if message.severity.upper() == "ERROR")
    warnings = sum(1 for message in bundle.validation_log if message.severity.upper() == "WARNING")
    return MarketDataSessionSummary(
        session_id=session_id,
        files=files,
        missing_required_files=missing_required_files,
        blocking_errors=blocking_errors,
        warnings=warnings,
        ready=not missing_required_files and blocking_errors == 0,
        validation_log=bundle.validation_log,
        counts=_counts_from_bundle(bundle),
        available_fx_pairs=_available_fx_pairs_from_bundle(bundle),
    )


def store_market_data_file(session_id: str, filename: str, content: bytes) -> MarketDataSessionSummary:
    safe_name = Path(filename).name
    kind = classify_market_data_filename(safe_name)
    if kind is None:
        raise ValueError(
            "Файл не распознан как market data bundle. Поддерживаются curveDiscount, curveForward, fixing, calibrationInstrument*, RC_*. "
        )
    _validate_xlsx_row_limit(safe_name, content)

    session_dir = _session_dir(session_id)
    session_dir.mkdir(parents=True, exist_ok=True)
    target = session_dir / safe_name
    target.write_bytes(content)
    return summarize_market_data_session(session_id)


def populate_market_data_session_from_directory(
    session_id: str,
    source_dir: Path,
    *,
    validate_row_limit: bool = True,
) -> MarketDataSessionSummary:
    if not source_dir.exists():
        raise ValueError(f"Каталог с market data не найден: {source_dir}")
    if not source_dir.is_dir():
        raise ValueError(f"Ожидается каталог с market data, получен файл: {source_dir}")

    session_dir = _session_dir(session_id)
    session_dir.mkdir(parents=True, exist_ok=True)
    for path in sorted(source_dir.iterdir()):
        if not path.is_file():
            continue
        if classify_market_data_filename(path.name) is None:
            continue
        if validate_row_limit:
            _validate_xlsx_row_limit(path.name, path)
        shutil.copy2(path, session_dir / path.name)
    return summarize_market_data_session(session_id)


def create_session_from_default_datasets() -> MarketDataSessionSummary:
    session_id = create_market_data_session()
    return populate_market_data_session_from_directory(
        session_id,
        _default_datasets_dir(),
        validate_row_limit=False,
    )


def populate_market_data_session_from_live_sources(
    session_id: str,
    *,
    as_of_date: date | None = None,
    lookback_days: int = 180,
) -> MarketDataSessionSummary:
    session_dir = _session_dir(session_id)
    session_dir.mkdir(parents=True, exist_ok=True)
    sync_live_market_data_to_directory(
        session_dir,
        as_of_date=as_of_date,
        lookback_days=lookback_days,
    )
    return summarize_market_data_session(session_id)


def create_session_from_live_sources(
    *,
    as_of_date: date | None = None,
    lookback_days: int = 180,
) -> MarketDataSessionSummary:
    session_id = create_market_data_session()
    return populate_market_data_session_from_live_sources(
        session_id,
        as_of_date=as_of_date,
        lookback_days=lookback_days,
    )


def load_market_data_bundle_for_session(session_id: str) -> tuple[MarketDataBundle, MarketDataSessionSummary]:
    summary = summarize_market_data_session(session_id)
    if summary.blocking_errors > 0 or summary.missing_required_files:
        missing = ", ".join(summary.missing_required_files) if summary.missing_required_files else "—"
        raise ValueError(
            f"Market data bundle не готов. Отсутствуют обязательные файлы: {missing}. "
            f"Блокирующих ошибок: {summary.blocking_errors}."
        )
    bundle = load_market_data_bundle_from_directory(_session_dir(session_id), strict=False)
    return bundle, summary


def read_market_data_session_metadata(session_id: str) -> dict:
    safe_id = validate_market_data_session_id(session_id)
    path = _session_dir(safe_id) / "marketDataMetadata.json"
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def find_latest_ready_market_data_session() -> MarketDataSessionSummary | None:
    root = _sessions_root()
    if not root.exists():
        return None

    ordered = sorted(
        [path for path in root.iterdir() if path.is_dir()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for path in ordered:
        try:
            summary = summarize_market_data_session(path.name)
        except (BadZipFile, InvalidFileException, KeyError) as exc:
            logger.warning(
                "skip_bad_market_data_session session_id=%s error_type=%s error=%s",
                path.name,
                type(exc).__name__,
                exc,
            )
            continue
        except Exception as exc:
            logger.warning(
                "skip_bad_market_data_session session_id=%s error_type=%s error=%s",
                path.name,
                type(exc).__name__,
                exc,
            )
            continue
        if summary.ready:
            return summary
    return None


def _empty_frame():
    import pandas as pd

    return pd.DataFrame()


__all__ = [
    "MarketDataSessionSummary",
    "UploadedMarketDataFile",
    "classify_market_data_filename",
    "clear_market_data_session",
    "create_market_data_session",
    "create_session_from_default_datasets",
    "create_session_from_live_sources",
    "find_latest_ready_market_data_session",
    "get_market_data_session_dir",
    "load_market_data_bundle_for_session",
    "read_market_data_session_metadata",
    "populate_market_data_session_from_live_sources",
    "populate_market_data_session_from_directory",
    "store_market_data_file",
    "summarize_market_data_session",
    "validate_market_data_session_id",
    "validate_market_data_xlsx_row_limit",
]
