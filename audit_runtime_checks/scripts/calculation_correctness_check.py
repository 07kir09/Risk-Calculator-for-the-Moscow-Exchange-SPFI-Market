from __future__ import annotations

import json
import math
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
BACKEND_URL = os.environ.get("BACKEND_URL", "http://127.0.0.1:8000").rstrip("/")
FX_RATES = {"USD": 90.0, "EUR": 100.0}
MONEY_TOL = 0.01
REL_TOL = 0.0001


def post_json(path: str, payload: dict, timeout: float = 30) -> tuple[int, bytes]:
    request = urllib.request.Request(
        f"{BACKEND_URL}{path}",
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return response.status, response.read()
    except urllib.error.HTTPError as exc:
        return exc.code, exc.read()


def forward(position_id: str, currency: str, quantity: float, notional: float, spot: float, strike: float) -> dict:
    return {
        "instrument_type": "forward",
        "position_id": position_id,
        "option_type": "call",
        "style": "european",
        "quantity": quantity,
        "notional": notional,
        "underlying_symbol": position_id.upper(),
        "underlying_price": spot,
        "strike": strike,
        "volatility": 0.0,
        "maturity_date": "2027-01-01",
        "valuation_date": "2026-01-01",
        "risk_free_rate": 0.0,
        "dividend_yield": 0.0,
        "currency": currency,
        "liquidity_haircut": 0.0,
    }


def positions() -> list[dict]:
    return [
        forward("rub_long", "RUB", 2.0, 10.0, 100.0, 95.0),
        forward("usd_long", "USD", 1.0, 5.0, 50.0, 48.0),
        forward("eur_long", "EUR", 3.0, 2.0, 80.0, 75.0),
        forward("usd_short", "USD", -1.0, 4.0, 70.0, 72.0),
    ]


def scenarios() -> list[dict]:
    return [
        {"scenario_id": "base", "underlying_shift": 0.0, "volatility_shift": 0.0, "rate_shift": 0.0},
        {"scenario_id": "shock_up", "underlying_shift": 0.10, "volatility_shift": 0.0, "rate_shift": 0.0},
        {"scenario_id": "shock_down", "underlying_shift": -0.05, "volatility_shift": 0.0, "rate_shift": 0.0},
        {"scenario_id": "stress_worst", "underlying_shift": -0.20, "volatility_shift": 0.0, "rate_shift": 0.0},
        {"scenario_id": "stress_mild", "underlying_shift": 0.02, "volatility_shift": 0.0, "rate_shift": 0.0},
    ]


def payload() -> dict:
    return {
        "positions": positions(),
        "scenarios": scenarios(),
        "base_currency": "RUB",
        "fx_rates": FX_RATES,
        "alpha": 0.80,
        "horizon_days": 1,
        "mode": "demo",
        "calc_sensitivities": False,
        "calc_var_es": True,
        "calc_stress": True,
        "calc_margin_capital": True,
        "calc_correlations": False,
    }


def fx(currency: str) -> float:
    return 1.0 if currency == "RUB" else FX_RATES[currency]


def manual_value(position: dict, scenario: dict | None = None) -> float:
    shift = 0.0 if scenario is None else float(scenario["underlying_shift"])
    rate_shift = 0.0 if scenario is None else float(scenario["rate_shift"])
    tenor_days = 365
    discount = math.exp(-(float(position["risk_free_rate"]) + rate_shift) * tenor_days / 365.0)
    spot = float(position["underlying_price"]) * (1.0 + shift)
    local = float(position["quantity"]) * float(position["notional"]) * (spot - float(position["strike"])) * discount
    return local * fx(str(position["currency"]))


def manual_expected() -> dict:
    pos = positions()
    scen = scenarios()
    pnl_matrix = [[manual_value(position, scenario) - manual_value(position) for scenario in scen] for position in pos]
    pnl_distribution = [sum(row[index] for row in pnl_matrix) for index in range(len(scen))]
    sorted_pnls = sorted(pnl_distribution)
    tail_count = max(1, math.ceil(len(sorted_pnls) * 0.20 - 1e-12))
    tail = sorted_pnls[:tail_count]
    return {
        "base_value": sum(manual_value(position) for position in pos),
        "pnl_distribution": pnl_distribution,
        "var_hist": max(0.0, -tail[-1]),
        "es_hist": max(0.0, -sum(tail) / len(tail)),
        "variation_margin": 0.0,
    }


def close_enough(actual: float, expected: float) -> tuple[bool, float]:
    delta = abs(float(actual) - float(expected))
    tolerance = max(MONEY_TOL, abs(float(expected)) * REL_TOL)
    return delta <= tolerance, delta


def export_json_excel(metrics: dict) -> tuple[dict, dict]:
    REPORTS.mkdir(parents=True, exist_ok=True)
    export_payload = {
        "methodology_metadata": {
            "methodology_status": "preliminary",
            "limit_source": "draft_auto",
            "preliminary": True,
            "var_method": "scenario_quantile",
            "stress_source": "backend_calculated",
            "backend_calculated": True,
            "export_generated_at": "runtime-check",
        },
        "metrics": metrics,
    }
    json_path = REPORTS / "calculation_correctness_export.json"
    json_path.write_text(json.dumps(export_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    excel_values: dict[str, float] = {}
    try:
        from openpyxl import Workbook, load_workbook

        xlsx_path = REPORTS / "calculation_correctness_export.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics"
        ws.append(["metric", "value"])
        for key in ["base_value", "var_hist", "es_hist", "variation_margin"]:
            ws.append([key, metrics.get(key)])
        meta = wb.create_sheet("Methodology")
        for key, value in export_payload["methodology_metadata"].items():
            meta.append([key, value])
        wb.save(xlsx_path)
        loaded = load_workbook(xlsx_path, data_only=True)
        for key, value in loaded["Metrics"].iter_rows(min_row=2, values_only=True):
            excel_values[str(key)] = float(value)
    except Exception as exc:
        excel_values["_error"] = str(exc)

    return export_payload, excel_values


def row(name: str, manual: object, backend: object, delta: object, ok: bool) -> dict[str, object]:
    print(f"{'PASS' if ok else 'FAIL'} {name}: manual={manual} backend={backend} delta={delta}")
    return {"metric": name, "manual": manual, "backend": backend, "delta": delta, "ok": "PASS" if ok else "FAIL"}


def main() -> int:
    status, raw = post_json("/metrics", payload())
    if status != 200:
        print(f"FAIL /metrics returned {status}: {raw[:500].decode('utf-8', 'replace')}")
        return 1

    backend = json.loads(raw.decode("utf-8"))
    expected = manual_expected()
    rows: list[dict[str, object]] = []
    for key in ["base_value", "var_hist", "es_hist", "variation_margin"]:
        ok, delta = close_enough(backend[key], expected[key])
        rows.append(row(key, expected[key], backend[key], delta, ok))

    for index, expected_pnl in enumerate(expected["pnl_distribution"]):
        actual = backend["pnl_distribution"][index]
        ok, delta = close_enough(actual, expected_pnl)
        rows.append(row(f"pnl_distribution[{index}]", expected_pnl, actual, delta, ok))

    json_export, excel_export = export_json_excel(backend)
    for key in ["base_value", "var_hist", "es_hist", "variation_margin"]:
        json_value = json_export["metrics"][key]
        excel_value = excel_export.get(key)
        json_ok, json_delta = close_enough(json_value, backend[key])
        excel_ok, excel_delta = close_enough(excel_value, backend[key]) if isinstance(excel_value, float) else (False, excel_value)
        rows.append(row(f"json_export.{key}", backend[key], json_value, json_delta, json_ok))
        rows.append(row(f"excel_export.{key}", backend[key], excel_value, excel_delta, excel_ok))

    lines = [
        "# Calculation correctness runtime report",
        "",
        "| Метрика | Manual/Backend | Backend/Export | Delta | Pass/Fail |",
        "| --- | ---: | ---: | ---: | --- |",
    ]
    for item in rows:
        lines.append(f"| {item['metric']} | {item['manual']} | {item['backend']} | {item['delta']} | {item['ok']} |")
    REPORTS.mkdir(parents=True, exist_ok=True)
    (REPORTS / "calculation_correctness_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return 0 if all(item["ok"] == "PASS" for item in rows) else 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except urllib.error.URLError:
        print("FAIL backend is not reachable. Start with: PYTHONPATH=backend uvicorn option_risk.api:app --app-dir backend --port 8000")
        raise SystemExit(1)
    except Exception as exc:
        print(f"FAIL {type(exc).__name__}: {exc}", file=sys.stderr)
        raise SystemExit(1)
